import openpyxl
from datetime import datetime


class ExcelControl:
    def __init__(self, wb_file_name):
        self.workbook = self.open_workbook(wb_file_name)
        self.excel_file_name = wb_file_name
        self.row_keyword_pair = {}
        self.row_num = 1

    # open workbook
    def open_workbook(self, file_name):
        try:
            wb = openpyxl.load_workbook(file_name)
            self.workbook = wb
        except FileNotFoundError:
            print(f"{file_name} 파일을 찾지 못했음")
            return None
        return wb

    def control_status_check(self):
        if self.workbook == None:
            return False
        else:
            return True

    def access_point_check(self, row, column):
        if row <= 1 or column <= 0:
            return False
        else:
            return True

    # data example : [["data", 2], ["data2", 3]]
    # (row, column) ~ (row + data row len, column + data column len)
    def _write_excel(self, data, row, column):
        ws = self.workbook.active
        if not self.control_status_check():
            return False
        if not self.access_point_check(row, column):
            return False
        for i, row_data in enumerate(data):
            for j, cell_data in enumerate(row_data):
                ws.cell(row=row + i, column=column + j, value=cell_data)
        try:
            # 23.07.02 : 엑셀 파일 별도 파일 저장
            today = datetime.today().strftime('%Y-%m-%d')
            excel_write_path = './결과/통합순위_' + today + '.xlsx'
            self.workbook.save(excel_write_path)
        except Exception as e:
            print(e)
            return False
        return True

    def _read_excel(self, row, column):
        ws = self.workbook.active
        if not self.control_status_check():
            return None
        if not self.access_point_check(row, column):
            return None
        try:
            cell_value = ws.cell(row=row, column=column).value
        except Exception as e:
            print(e)
            return None
        return cell_value

    def is_ended(self, row_idx):
        # 24.02.08 : URL이 없어도 처리하는 기능 추가
        ret = self._read_excel(row_idx, 1)
        if ret is None or ret == "":
            return True
        return False

    def read_keyword(self):
        row_idx = 2
        col_idx = 2
        keywords = []
        while True:
            ret = self._read_excel(row_idx, col_idx)
            # 테이블이 끝났거나 keyword가 존재하지 않으면 break.
            if self.is_ended(row_idx):
                break
            if ret is None or ret == "":
                break
            # self.row_keyword_pair.append([row_idx, ret])
            # self.row_keyword_pair[ret] = row_idx
            keywords.append(ret)
            row_idx += 1

        return keywords

    def read_urls(self):
        row_idx = 2
        col_idx = 3
        urls = []
        while True:
            ret = self._read_excel(row_idx, col_idx)
            # 테이블이 끝나면 break.
            if self.is_ended(row_idx):
                break
            if ret is None or ret == "":
                ret = ""
            # self.row_keyword_pair[ret] = row_idx
            urls.append(ret)
            row_idx += 1
        self.row_num = 1
        return urls

    def read_all(self):
        keywords = self.read_keyword()
        urls = self.read_urls()
        return keywords, urls

    def write_rank(self, rank, rank_col=4):
        self.row_num += 1
        if rank == -1:
            ret = self._write_excel([['']], self.row_num, rank_col)
            return ret
        # row = self.row_keyword_pair[url]
        if not self.access_point_check(self.row_num, rank_col):
            return False
        ret = self._write_excel([[rank]], self.row_num, rank_col)
        return ret


if __name__ == '__main__':
    ec = ExcelControl("rank_check.xlsx")
    if ec._write_excel([[3]], 2, 4):
        print("success")
    else:
        print("failed")
    ret = ec._read_excel(2, 3)
    print(ret)
    ret = ec._read_excel(2, 4)
    print(ret)
